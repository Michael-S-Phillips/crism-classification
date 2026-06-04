"""Harvest extra plag positive patches from sensor-space ROIs in the xlsx.

Source: previous_and_new_plag.xlsx, tab "Search through delivery 004".

Each plag ROI gives:
  obsid (e.g. FRT000064B3_07), numerator_row, numerator_col, roi_size_row, roi_size_col

The row/col are in MTRDR **detector (sensor) space**. To get the corresponding
pixels in the projected mtr3 cube we open the paired ``*_in*_mtr3.img`` file —
band 7 is IR Line (detector row per projected pixel), band 6 is IR Sample
(detector col). Pixels whose (IR Line, IR Sample) fall inside the ROI box are
the in-ROI projected pixels.

For each in-ROI projected pixel we read a 7x7x489 window from the if cube,
resample each spectrum to the 59 MRDR wavelengths, clip [0, 0.5], and emit
``(N, 7, 7, 59)`` float32 + a meta.parquet matching the positives-pool schema.

Output schema is intentionally identical to ``scripts/build_contrastive_data.py``
positives, so the result can be merged directly via
``--extra_positive_pool_dir``.

Usage:

  conda run -n crism python scripts/build_plag_roi_patches.py \\
      --xlsx_path previous_and_new_plag.xlsx \\
      --output_dir data/contrastive/extra_plag_roi
"""
from __future__ import annotations

import argparse
import glob
import logging
import os
import re
import sys
from typing import List, Optional, Tuple

import numpy as np
import openpyxl
import pandas as pd
import rasterio
import spectral.io.envi as envi

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from data.synthetic_plag import interp_to_mrral_wavelengths

# ----------------------------------------------------------------- constants
NODATA = 65535.0
CLIP_MAX = 0.5
PATCH_SIZE = 7
N_BANDS = 59
PAD = PATCH_SIZE // 2

FELDSREVIEW_ROOT_DEFAULT = '/mnt/mrdr/categorized_mineral_units/FeldsReview'
EXCLUDE_MARKER = 'ACTUALLY NOT GOOD'

# In-file band order (1-indexed for rasterio):
#   1 VNIR/IR Spectral Continuity Residual
#   2 VNIR/IR Spatial Gradient Residual
#   3 ATM Correction Spectral Shift Artifact
#   4 VNIR Sample, 5 VNIR Line
#   6 IR Sample,   7 IR Line
#   8 VNIR/IR Ground Sampling Offset, 9 VNIR/IR Mask
IR_SAMPLE_BAND = 6
IR_LINE_BAND = 7

logger = logging.getLogger(__name__)


# ----------------------------------------------------------------- xlsx parser
def _find_subheader_row(raw: list[tuple]) -> Optional[int]:
    """Find the row index containing the three ROW/COL subheader pairs."""
    for i, r in enumerate(raw):
        if r is None:
            continue
        cells = [str(c).strip().lower() if c else '' for c in r]
        if cells.count('row') >= 3 and cells.count('col') >= 3:
            return i
    return None


def _parse_columns(raw: list[tuple], subheader_i: int) -> Optional[dict]:
    """Return positions of {region, id, sig, conf, num_row, num_col,
    roi_row, roi_col}.

    Both xlsx layouts share three ``ROW/COL`` subheader pairs (Numerator,
    Denominator, ROI size) in the same row. ``Search through delivery 004``
    has explicit ``Region`` / ``ID`` / ``Ratioed spectral signature`` /
    ``Confidence`` text headers in the same subheader row; ``Plag from
    literature`` does not, and we fall back to hard-coded column positions
    (col 0 = region, 1 = id, 2 = confidence, 6 = signature) verified against
    that sheet's data rows.
    """
    h = [str(c).strip().lower() if c else '' for c in raw[subheader_i]]
    row_indices = [i for i, c in enumerate(h) if c == 'row']
    col_indices = [i for i, c in enumerate(h) if c == 'col']
    if len(row_indices) < 3 or len(col_indices) < 3:
        return None
    cols = {
        'num_row': row_indices[0],
        'num_col': col_indices[0],
        'roi_row': row_indices[2],
        'roi_col': col_indices[2],
    }
    # Try header-based first (Search through delivery 004 layout)
    cols['region'] = h.index('region') if 'region' in h else None
    cols['id'] = h.index('id') if 'id' in h else None
    cols['sig'] = next((i for i, c in enumerate(h)
                        if 'ratioed' in c or 'signature' in c), None)
    cols['conf'] = next((i for i, c in enumerate(h)
                         if 'confidence' in c), None)
    if cols['region'] is None or cols['id'] is None:
        # Plag from literature: no text headers in subheader row; use
        # hard-coded positions verified against actual data rows.
        cols['region'] = 0
        cols['id'] = 1
        cols['conf'] = 2
        cols['sig'] = 6
    return cols


def _clean_obsid(raw) -> Optional[str]:
    """Normalize an obsid cell: strip whitespace/non-breaking, remove '_07' /
    '_09' suffix, lowercase. Returns None if not a valid CRISM obsid pattern."""
    if not isinstance(raw, str):
        return None
    s = raw.strip().rstrip('\xa0').lower()
    s = re.sub(r'_0[79].*$', '', s)
    if not re.match(r'^(frt|hrl|hrs|frs)[0-9a-f]+$', s):
        return None
    return s


def parse_roi_rows(xlsx_path: str,
                   include_low: bool = False,
                   include_non_plag: bool = False,
                   sheets: Optional[list[str]] = None) -> pd.DataFrame:
    """Return a DataFrame of plag ROI rows from one or more xlsx sheets.

    Columns: sheet, region, obsid, signature, confidence,
             num_row, num_col, roi_row, roi_col.
    Deduplicated by (obsid, num_row, num_col, roi_row, roi_col).
    """
    wb = openpyxl.load_workbook(xlsx_path)
    sheets_to_read = sheets if sheets is not None else list(wb.sheetnames)
    out_rows: list[dict] = []
    for sn in sheets_to_read:
        if sn not in wb.sheetnames:
            logger.warning(f'sheet not found: {sn!r}')
            continue
        ws = wb[sn]
        raw = list(ws.iter_rows(values_only=True))
        header_i = _find_subheader_row(raw)
        if header_i is None:
            logger.info(f'  sheet {sn!r}: no ROW/COL subheader row found, skipping')
            continue
        cols = _parse_columns(raw, header_i)
        if cols is None:
            logger.warning(f'  sheet {sn!r}: could not parse columns, skipping')
            continue

        last_region = None
        last_obsid = None
        in_exclude = False
        kept_in_sheet = 0
        for r in raw[header_i + 1:]:
            if r is None or all(c is None for c in r):
                continue
            first = r[0]
            if isinstance(first, str) and EXCLUDE_MARKER in first:
                in_exclude = True
                continue
            if in_exclude:
                continue
            # Skip subsection / legend rows (e.g. starting with '*' or
            # 'Found in literature')
            if isinstance(first, str) and (
                first.startswith('*') or 'found in literature' in first.lower()
            ):
                continue
            region = r[cols['region']] if r[cols['region']] is not None else last_region
            last_region = region or last_region
            obsid_raw = r[cols['id']] if r[cols['id']] is not None else last_obsid
            last_obsid = obsid_raw or last_obsid
            obsid = _clean_obsid(obsid_raw)
            if obsid is None:
                continue
            signature = r[cols['sig']] if cols['sig'] is not None else None
            confidence = r[cols['conf']] if cols['conf'] is not None else None
            num_row = r[cols['num_row']]
            num_col = r[cols['num_col']]
            roi_row = r[cols['roi_row']]
            roi_col = r[cols['roi_col']]
            if not all(isinstance(x, (int, float))
                       for x in (num_row, num_col, roi_row, roi_col)):
                # "No MTRDR" / "nd" rows have '-' for numeric cols
                continue
            out_rows.append({
                'sheet': sn,
                'region': region or '',
                'obsid': obsid,
                'signature': str(signature).strip() if signature else '',
                'confidence': str(confidence).strip() if confidence else '',
                'num_row': int(num_row),
                'num_col': int(num_col),
                'roi_row': int(roi_row),
                'roi_col': int(roi_col),
            })
            kept_in_sheet += 1
        logger.info(f'  sheet {sn!r}: kept {kept_in_sheet} candidate rows')

    df = pd.DataFrame(out_rows)
    if df.empty:
        return df
    if not include_non_plag:
        df = df[df['signature'].str.lower().str.contains('plagioclase')]
    if not include_low:
        df = df[~df['confidence'].str.lower().str.startswith('low')]
    # Dedup across sheets (same ROI may appear in both tabs)
    df = df.drop_duplicates(
        subset=['obsid', 'num_row', 'num_col', 'roi_row', 'roi_col']
    ).reset_index(drop=True)
    return df


# ----------------------------------------------------------------- file lookup
def build_inventory(root: str) -> dict:
    """One-shot recursive walk of ``root`` for ``*_if*_mtr3.img`` and
    ``*_in*_mtr3.img`` files. Returns ``{obsid: {'if': [paths], 'in': [paths]}}``.

    Avoids per-obsid recursive globs (which are O(N_obsids * tree_size) on the
    network share). Single walk is ~one minute even on a slow mount.
    """
    paths = glob.glob(os.path.join(root, '**', '*_mtr3.img'), recursive=True)
    inv: dict = {}
    for p in paths:
        fname = os.path.basename(p).lower()
        m = re.match(r'^((?:frt|hrl|hrs|frs)[0-9a-f]+)', fname)
        if not m:
            continue
        oid = m.group(1)
        kind = 'if' if '_if' in fname else 'in' if '_in' in fname else None
        if kind is None:
            continue
        inv.setdefault(oid, {'if': [], 'in': []})[kind].append(p)
    return inv


def find_pair_for_obsid(
    obsid: str,
    root: str,
    inventory: Optional[dict] = None,
) -> Optional[Tuple[str, str]]:
    """Return ``(if_path, in_path)`` if both exist for this obsid, else None.

    Prefers ``inventory`` (built once via :func:`build_inventory`) for O(1)
    lookups. Falls back to per-obsid recursive globs if no inventory is
    provided.
    """
    obsid = obsid.lower()
    if inventory is not None:
        info = inventory.get(obsid) or {}
        ifs = info.get('if', [])
        ins = info.get('in', [])
        if ifs and ins:
            # Prefer pairs from the same directory (sibling files).
            if_path = ifs[0]
            sib = os.path.dirname(if_path)
            sib_in = [p for p in ins if os.path.dirname(p) == sib]
            in_path = sib_in[0] if sib_in else ins[0]
            return if_path, in_path
        return None
    short = re.sub(r'^frt0+', 'frt', obsid)  # 'frt000064b3' -> 'frt64b3' fallback
    for tag in (obsid, short):
        if_candidates = glob.glob(os.path.join(
            root, '**', f'{tag}*_if*_mtr3.img'), recursive=True)
        in_candidates = glob.glob(os.path.join(
            root, '**', f'{tag}*_in*_mtr3.img'), recursive=True)
        if not if_candidates or not in_candidates:
            continue
        if_path = if_candidates[0]
        sib_dir = os.path.dirname(if_path)
        sib_in = [p for p in in_candidates if os.path.dirname(p) == sib_dir]
        if sib_in:
            return if_path, sib_in[0]
    return None


# ----------------------------------------------------------------- harvest
def harvest_roi_patches(
    if_path: str,
    in_path: str,
    num_row: int,
    num_col: int,
    roi_row: int,
    roi_col: int,
    target_wl: np.ndarray,
) -> List[Tuple[np.ndarray, int, int]]:
    """Return ``[(patch_7x7x59, proj_row, proj_col), ...]`` for in-ROI pixels.

    Performance note: we never read the full if cube. We compute the in-ROI
    projected pixel set first via the in file (read just 2 bands of that),
    then read only the bounding-box window around those pixels from the if
    cube — at most ~16x16 spatial for a 5x5 sensor ROI, instead of the full
    669x748x489 ~245MB cube.
    """
    with rasterio.open(in_path) as ds_in:
        ir_line = ds_in.read(IR_LINE_BAND).astype(np.int32)
        ir_sample = ds_in.read(IR_SAMPLE_BAND).astype(np.int32)
    H, W = ir_line.shape

    rh = roi_row // 2
    rw = roi_col // 2
    mask = (
        (ir_line   >= num_row - rh) & (ir_line   <= num_row + rh) &
        (ir_sample >= num_col - rw) & (ir_sample <= num_col + rw) &
        (ir_line   < int(NODATA))   & (ir_sample < int(NODATA))
    )
    rr, cc = np.where(mask)
    if rr.size == 0:
        return []

    cube_wl = np.asarray(
        envi.open(if_path.replace('.img', '.hdr')).bands.centers,
        dtype=np.float64,
    )

    # Read ONLY the bounding-box region of the ROI (with PAD for 7x7 patches)
    r_min = max(0, int(rr.min()) - PAD)
    r_max = min(H, int(rr.max()) + PAD + 1)
    c_min = max(0, int(cc.min()) - PAD)
    c_max = min(W, int(cc.max()) + PAD + 1)
    with rasterio.open(if_path) as ds_if:
        win = rasterio.windows.Window(
            c_min, r_min, c_max - c_min, r_max - r_min)
        sub = ds_if.read(window=win).astype(np.float32)        # (B, h, w)
    sub = np.where(sub == NODATA, np.nan, sub)

    out: List[Tuple[np.ndarray, int, int]] = []
    for r, c in zip(rr, cc):
        # window-local center
        lr = r - r_min
        lc = c - c_min
        # 7x7 neighbourhood, all within the sub region by construction
        r0 = max(0, lr - PAD); r1 = min(sub.shape[1], lr + PAD + 1)
        c0 = max(0, lc - PAD); c1 = min(sub.shape[2], lc + PAD + 1)
        chunk = sub[:, r0:r1, c0:c1].transpose(1, 2, 0)        # (h, w, B)
        cr = lr - r0
        cc_local = lc - c0
        center_spec = chunk[cr, cc_local]
        if not np.isfinite(center_spec).any():
            continue
        patch = np.zeros((PATCH_SIZE, PATCH_SIZE, N_BANDS), dtype=np.float32)
        dr0 = max(0, PAD - lr)
        dc0 = max(0, PAD - lc)
        for pi in range(chunk.shape[0]):
            for pj in range(chunk.shape[1]):
                spec = chunk[pi, pj]
                if np.isfinite(spec).sum() < 5:
                    continue   # leave zeros — edge / all-NODATA neighbour
                resampled = interp_to_mrral_wavelengths(
                    cube_wl, spec, target_wl).astype(np.float32)
                patch[dr0 + pi, dc0 + pj] = resampled
        patch = np.clip(patch, 0.0, CLIP_MAX)
        out.append((patch, int(r), int(c)))
    return out


# ----------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx_path', default='previous_and_new_plag.xlsx')
    ap.add_argument('--feldsreview_root', default=FELDSREVIEW_ROOT_DEFAULT)
    ap.add_argument('--mrral_hdr', default=None,
                    help='Any mrral .hdr to read the 59 MRDR target wavelengths from. '
                         'Default: first /mnt/mrdr/mc*/t*_mrral_*.hdr found.')
    ap.add_argument('--output_dir', default='data/contrastive/extra_plag_roi')
    ap.add_argument('--include_low', action='store_true',
                    help='Also harvest rows whose confidence starts with "Low".')
    ap.add_argument('--include_non_plag', action='store_true',
                    help='Also harvest rows whose signature is not plagioclase.')
    ap.add_argument('-v', '--verbose', action='store_true')
    args = ap.parse_args()
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format='%(asctime)s %(levelname)s %(message)s',
    )

    mrral_hdr = args.mrral_hdr or sorted(
        glob.glob('/mnt/mrdr/mc*/t*_mrral_*.hdr'))[0]
    target_wl = np.asarray(
        envi.open(mrral_hdr).bands.centers, dtype=np.float64,
    )[:N_BANDS]
    logger.info(f'target MRDR wavelengths: {len(target_wl)} bands, '
                f'{target_wl[0]:.1f}-{target_wl[-1]:.1f} nm')

    df = parse_roi_rows(args.xlsx_path,
                        include_low=args.include_low,
                        include_non_plag=args.include_non_plag)
    logger.info(f'parsed {len(df)} plag ROI rows from {args.xlsx_path}')
    logger.info(f'building one-shot inventory under {args.feldsreview_root} ...')
    inventory = build_inventory(args.feldsreview_root)
    logger.info(f'  inventory: {len(inventory)} obsids with mtr3 files on disk')

    all_patches: list[np.ndarray] = []
    all_rows: list[dict] = []
    n_skipped_missing = 0
    for idx, row in df.iterrows():
        obsid = row['obsid']
        pair = find_pair_for_obsid(obsid, args.feldsreview_root,
                                   inventory=inventory)
        if pair is None:
            logger.warning(f'  [{obsid}] if/in pair not found, skipping')
            n_skipped_missing += 1
            continue
        if_path, in_path = pair
        logger.info(f'  [{obsid}] ROI ({row.num_row},{row.num_col}) size '
                    f'{row.roi_row}x{row.roi_col} -> {os.path.basename(if_path)}')
        try:
            patches = harvest_roi_patches(
                if_path, in_path,
                row['num_row'], row['num_col'],
                row['roi_row'], row['roi_col'], target_wl,
            )
        except Exception as e:
            logger.warning(f'    error on {obsid} row {idx}: {e}')
            continue
        for p, r, c in patches:
            all_patches.append(p)
            all_rows.append({
                'pixel_row': r,
                'pixel_col': c,
                'tile_id': obsid,
                'source_polygon': f'roi_{idx}',
                'source_gpkg': 'previous_and_new_plag.xlsx',
                'confidence': row['confidence'],
                'signature': row['signature'],
                'region': row['region'],
            })
        logger.info(f'    yielded {len(patches)} patches '
                    f'(running total: {len(all_patches)})')

    os.makedirs(args.output_dir, exist_ok=True)
    if all_patches:
        arr = np.stack(all_patches, axis=0).astype(np.float32)
    else:
        arr = np.zeros((0, PATCH_SIZE, PATCH_SIZE, N_BANDS), dtype=np.float32)
    np.save(os.path.join(args.output_dir, 'patches.npy'), arr)
    pd.DataFrame.from_records(all_rows).to_parquet(
        os.path.join(args.output_dir, 'meta.parquet'), index=False)
    logger.info(f'wrote {len(arr)} patches + meta to {args.output_dir}')
    logger.info(f'  obsids skipped (if/in pair missing on disk): {n_skipped_missing}')


if __name__ == '__main__':
    main()
